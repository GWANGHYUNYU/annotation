import os
import cv2
import numpy as np
from lxml import etree
import xml.etree.cElementTree as ET
import openpyxl

from file_folder_analysis_dialog import FileFolderAnalysisDialog

from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *


class FileManager(QWidget):
    def __init__(self, parent=None):
        super(FileManager, self).__init__(parent)

        # 부모인 메인 프로그램 클래스 변수/함수 상속
        self.parent = parent

        self.VALID_FORMAT = ('.BMP', '.GIF', '.JPG', '.JPEG', '.PNG', '.PBM', '.PGM', '.PPM', '.TIFF',
                             '.XBM')  # Image formats supported by Qt

        self.ELE_LIST = ['Head', 'Wire', 'entire']
        self.n_objects = len(self.ELE_LIST)



        # 파일 분석 관련 변수
        self.class_list = []
        self.class_cnt_list = []
        self.species_list = []
        self.species_cnt_list = []
        self.species_class_index_list = []

        self.species_ele_cnt_list = [ [] for _ in range(self.n_objects) ]
        self.class_ele_cnt_list =[ [] for _ in range(self.n_objects) ]
      

        self.total_ele_cnt_list = [0] * self.n_objects
        self.total_cnt = 0

       
        self.file_folder_analysis_dialog = FileFolderAnalysisDialog(parent=self)

    # 영상 파일에 해당되는 경로만 처리
    def check_is_images(self, path):
        if path.upper().endswith(self.VALID_FORMAT):
            isfile = True
        else:
            isfile = False

        return isfile

    # 입력 경로(한글)에 대한 영상 파일 열기
    def hangulFilePathImageRead(self, filePath):
        # 선택한 경로가 폴더가 아니고 파일인 경우만 영상열기 수행, 아닌경우 예외처리
        if os.path.isfile(filePath):
            # 파일경로 확장자 확인하여 영상 파일인지 아닌지 확인
            isfile = self.check_is_images(filePath)

            # 파일경로 확장자가 영상 파일인 경우 영상 디코딩 수행하여 영상 열기 수행
            if isfile is True:
                stream = open(filePath.encode("utf-8"), "rb")
                bytes = bytearray(stream.read())
                numpyArray = np.asarray(bytes, dtype=np.uint8)

                # 파일 디코딩 수행하여 영상 열기 성공한 경우에만 영상 리턴, 아닌경우 예외처리
                try:
                    self.parent.flg_img_ok = True
                    self.parent.input_img = cv2.imdecode(numpyArray, cv2.IMREAD_UNCHANGED)
                    self.parent.proc_img = self.parent.input_img.copy()
                except:
                    self.parent.flg_img_ok = False
                    self.parent.input_img = None
                    self.parent.proc_img = None
            # 파일이 영상파일이 아닌경우 스킵
            else:
                self.parent.flg_img_ok = False
                self.parent.input_img = None
                self.parent.proc_img = None
        # 파일이 아니고 폴더인 경우 스킵
        else:
            self.parent.flg_img_ok = False
            self.parent.input_img = None
            self.parent.proc_img = None

    def file_folder_analysis(self, sel_path):
        file_all_count = 0

        # 선택폴더의 전체 영상 개수 확인
        for root, dirs, files in os.walk(sel_path):
            for fname in files:
                if fname.upper().endswith(self.VALID_FORMAT):
                    file_all_count = file_all_count + 1

        # 처리과정 보이기 위해 프로그레스바 설정
        self.parent.progressBar_main.setRange(0, file_all_count)

        # 계산 결과 저장용 리스트 변수들 초기화

        self.class_list = []
        self.class_cnt_list = []
        self.species_list = []
        self.species_cnt_list = []
        self.species_class_index_list = []

        self.species_ele_cnt_list = [ [] for _ in range(self.n_objects) ]
        self.class_ele_cnt_list =[ [] for _ in range(self.n_objects) ]
        self.total_ele_cnt_list = [0] * self.n_objects
        self.total_cnt = 0

       
        process_cnt = 0
        # 선택폴더에 대해서 파일 분석 수행
        for root, dirs, files in os.walk(sel_path):
            # 모든 파일에 대해서 수행
            for fname in files:
                # 파일이 영상파일인 경우에 대해서만 수행
                if fname.upper().endswith(self.VALID_FORMAT):
                    # 파일 전체 경로 확인
                    full_path = os.path.join(root, fname)
                    # 상위폴더를 식물종으로 사용하고 해당 경로와 이름 확인
                    species_path = os.path.dirname(full_path)
                    species_name = os.path.basename(species_path)
                    # 상상위폴더를 식물과로 사용하고 해당 경로와 이름 확인
                    class_path = os.path.dirname(species_path)
                    class_name = os.path.basename(class_path)

                    # 확인된 식물과 정보가 기존 리스트에 없으면 정보 추가
                    if class_name not in self.class_list:
                        self.class_list.append(class_name)
                        self.class_cnt_list.append(1)

                        # 어노테이션 클래스 개수 세기용 리스트 초기화
                        for n in range(self.n_objects):
                            self.class_ele_cnt_list[n].append(0)
                       

                    # 확인된 식물과 정보가 기존리스트에 있으면 해당 식물과 개수 증가
                    else:
                        index_class = self.class_list.index(class_name)
                        self.class_cnt_list[index_class] += 1

                    # 확인된 식물종 정보가 기존 리스트에 없으면 정보 추가
                    if species_name not in self.species_list:
                        self.species_list.append(species_name)
                        self.species_cnt_list.append(1)
                        self.species_class_index_list.append(self.class_list.index(class_name))

                        # 어노테이션 클래스 개수 세기용 리스트 초기화
                        for n in range(self.n_objects):
                            self.species_ele_cnt_list[n].append(0)
                       

                    # 확인된 식물종 정보가 기존리스트에 있으면 해당 식물종 개수 증가
                    else:
                        index_species = self.species_list.index(species_name)
                        self.species_cnt_list[index_species] += 1

                    # 해당 영상에 대해서 어노테이션 파일 분석 수행
                    input_dir = os.path.dirname(full_path)
                    base_folder = os.path.dirname(input_dir)
                    output_dir = os.path.join(base_folder, os.path.basename(input_dir) + '_annotated')
                    annotation_paths = self.parent.annot_manager.get_annotation_paths(full_path,
                                                                                      self.parent.annot_manager.annotation_formats,
                                                                                      input_dir, output_dir)
                    ann_path = next(path for path in annotation_paths if 'PASCAL_VOC' in path)
                    if os.path.isfile(ann_path):
                        # 어노테이션 데이트 파일 경로 확인 및 정보 로드
                        tree = ET.parse(ann_path)
                        annotation = tree.getroot()
                        # 어노테이션 정보에서 각 정보들 로드
                        object_list = annotation.findall('object')
                        for obj in object_list:
                            index_species = self.species_list.index(species_name)
                            index_class = self.class_list.index(class_name)
                            anno_class_name, class_index, xmin, ymin, xmax, ymax = self.parent.annot_manager.get_xml_object_data(obj)

                            # 식물종에 대한 어노테이션 클래스별 개수 증가
                            # print(anno_class_name)
                            
                            for i in range(self.n_objects):

                                if anno_class_name == self.ELE_LIST[i]:
                                   
                                    self.species_ele_cnt_list[i][index_species] += 1
                                    self.class_ele_cnt_list[i][index_class] += 1
                                  
                    # 파일 처리 진행률 확인
                    process_cnt += 1
                    self.parent.progressBar_main.setValue(process_cnt)

        for i in range(0, len(self.class_list)):
            self.total_cnt += self.class_cnt_list[i]
            for n in range(self.n_objects):
                self.total_ele_cnt_list[n] += self.class_ele_cnt_list[n][i]


        # 결과 출력 다이얼로그 띄우기
        ret = self.file_folder_analysis_dialog.showModal()
        if ret:
            xl_filename = 'file_folder_analysis_result.xlsx'
            xl_book = openpyxl.Workbook()
            xl_sheet = xl_book.active

            xl_row_cnt = 1

            xl_sheet.cell(xl_row_cnt, 1, '식물과')
            xl_sheet.cell(xl_row_cnt, 2, '식물종')
            xl_sheet.cell(xl_row_cnt, 3, '원본영상')
            for i in range(self.n_objects):
                xl_sheet.cell(xl_row_cnt, 4 + i, self.ELE_LIST[i])
         

            xl_row_cnt += 1

            for i in range(0, len(self.class_list)):
                class_str = str(self.class_list[i])
                species_str = str(self.class_list[i])
                species_cnt_str = self.class_cnt_list[i]
              
                ele_cnt_str  = [str(self.class_ele_cnt_list[n][i]) for n in range(self.n_objects)]  
              
                item_list = [class_str, species_str, species_cnt_str] + ele_cnt_str

                # 해당 어노테이션 정보에 대해서 부모 클래스의 어노테이션 디스플레이 테이블에 정보 추가
               

                for column in range(len(item_list)):
                    xl_sheet.cell(xl_row_cnt, column + 1, item_list[column])

                xl_row_cnt += 1
            
            item_list = ['합계', '합계', self.total_cnt] + self.total_ele_cnt_list


            for column in range(len(item_list)):
                xl_sheet.cell(xl_row_cnt, column + 1, item_list[column])
            xl_row_cnt += 1


            for i in range(1, self.n_objects + 4):
                xl_sheet.cell(xl_row_cnt, i, '')
            xl_row_cnt += 1

            for i in range(0, len(self.species_list)):
                class_str = str(self.class_list[self.species_class_index_list[i]])
                species_str = str(self.species_list[i])
                species_cnt_str = self.species_cnt_list[i]
                ele_cnt_str  = [str(self.species_ele_cnt_list[n][i]) for n in range(self.n_objects)] 
             
                # 해당 어노테이션 정보에 대해서 부모 클래스의 어노테이션 디스플레이 테이블에 정보 추가
                item_list = [class_str, species_str, species_cnt_str] + ele_cnt_str
               
                for column in range(len(item_list)):
                    xl_sheet.cell(xl_row_cnt, column + 1, item_list[column])

                xl_row_cnt += 1

            xl_book.save(xl_filename)
